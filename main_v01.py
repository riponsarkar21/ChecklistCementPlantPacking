import tkinter as tk
from tkinter import messagebox
from tkcalendar import Calendar
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Path for the questions Excel file and the result file
questions_file = 'files/questions.xlsx'
result_file = 'files/submissions.xlsx'

# Load the questions sheet
def load_questions():
    wb = openpyxl.load_workbook(questions_file)
    sheet = wb['question']
    questions = []
    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
        label = sheet.cell(row=row, column=3).value  # Column C for the label
        if label:
            questions.append(label)
    return questions

# Create or load the result Excel file
def load_or_create_result_file():
    try:
        wb = openpyxl.load_workbook(result_file)
    except FileNotFoundError:
        wb = Workbook()
        wb.save(result_file)
    return wb

# Save the data to the result Excel file
def save_data(date, shift, data):
    wb = load_or_create_result_file()
    sheet = wb.active
    header = ['Date', 'Shift', 'Visitor Name', 'Label']
    
    # Check if header exists, if not, create it
    if sheet.max_row == 1:
        for col, h in enumerate(header, start=1):
            sheet.cell(row=1, column=col, value=h)

    # Check if the date and shift already exist, and overwrite the row
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == date and sheet.cell(row=row, column=2).value == shift:
            for col, value in enumerate(data, start=1):
                sheet.cell(row=row, column=col+2, value=value)  # Skip 'Date' and 'Shift' columns
            wb.save(result_file)
            return

    # If not found, append new row
    new_row = [date, shift] + data
    sheet.append(new_row)
    wb.save(result_file)

# Submit form data
def submit():
    date = cal.get_date()
    shift = shift_var.get()
    
    if not date or not shift:
        messagebox.showwarning("Input Error", "Please fill in all fields before submitting.")
        return

    data = []
    for var, entry in zip(checkbox_vars, visitor_name_entries):
        if var.get():
            data.append((entry.get() if entry.get() else ""))
        else:
            data.append("")
    
    # Save the data
    save_data(date, shift, data)
    messagebox.showinfo("Success", "Data submitted successfully!")

# GUI Setup
root = tk.Tk()
root.title("Visitor Data Form")

questions = load_questions()

# Create the form elements
checkbox_vars = []
visitor_name_entries = []
for idx, question in enumerate(questions, start=1):
    frame = tk.Frame(root)
    label = tk.Label(frame, text=question)
    label.grid(row=idx, column=0)

    var = tk.BooleanVar()
    checkbox = tk.Checkbutton(frame, variable=var)
    checkbox.grid(row=idx, column=1)

    entry = tk.Entry(frame)
    entry.grid(row=idx, column=2)
    
    checkbox_vars.append(var)
    visitor_name_entries.append(entry)
    frame.grid(row=idx, column=0, sticky="w")

# Date Picker
cal_label = tk.Label(root, text="Select Date:")
cal_label.grid(row=len(questions)+1, column=0)
cal = Calendar(root, selectmode="day", date_pattern="yyyy-mm-dd")
cal.grid(row=len(questions)+1, column=1)

# Shift Picker
shift_label = tk.Label(root, text="Select Shift:")
shift_label.grid(row=len(questions)+2, column=0)
shift_var = tk.StringVar()
shift_dropdown = tk.OptionMenu(root, shift_var, "A", "B", "C")
shift_var.set("A")  # Default shift
shift_dropdown.grid(row=len(questions)+2, column=1)

# Submit Button
submit_button = tk.Button(root, text="Submit", command=submit)
submit_button.grid(row=len(questions)+3, column=0, columnspan=2)

root.mainloop()
